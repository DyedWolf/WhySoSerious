from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, HttpResponse
from django.core.files.uploadedfile import InMemoryUploadedFile

from app01 import models
from app01.utils.pagination import Pagination
from app01.utils.form import DepartModelForm


def depart_list(request):
    """部门列表"""
    data_dict = {}
    title = request.GET.get("title", "")
    if title:
        data_dict["title__contains"] = title
    queryset = models.Department.objects.filter(**data_dict).order_by("id")
    page_object = Pagination(request, queryset)
    context = {"data_list": page_object.page_queryset, "page_string": page_object.html(), "search_data": title}
    return render(request, "depart_list.html", context)


def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        form = DepartModelForm()
        return render(request, "change.html", {"form": form, "title": "添加部门",
                                               "way": "depart"})
    form = DepartModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect("/depart/list/")
    context = {"form": form,
               "title": "添加部门",
               "way": "depart"}
    return render(request, "change.html", context)


def depart_delete(request):
    """删除部门"""
    nid = request.GET.get("nid")
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """编辑部门"""
    title = models.Department.objects.filter(id=nid).first()
    # print(title.id, title.title)
    if request.method == "GET":
        return render(request, "depart_edit.html", {"title": title.title})
    if request.method == "POST":
        new_title = request.POST.get("title")
        models.Department.objects.filter(id=nid).update(title=new_title)
        return redirect("/depart/list/")


@csrf_exempt
def depart_multi(request):
    """ 批量删除 """
    # 获取用户上传文件对象
    file_object = request.FILES.get("exc")
    if not file_object:
        return HttpResponse("未上传文件")
    # 对象传递给openpyxl ，由openpyxl 读取文件内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 循环获取
    # cell = sheet.cell(1, 1)
    # print(cell.value)
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exits = models.Department.objects.filter(title=text).exists()
        if not exits:
            models.Department.objects.create(title=text)
            print(text)
    return redirect("/depart/list/")
